import openpyxl, json, os

wb = openpyxl.load_workbook(r'c:\Users\user\Desktop\AGRI_AFRIQUE\AfricaAgri_Database_2000_2024.xlsx', data_only=True)
output_dir = r'c:\Users\user\Desktop\AGRI_AFRIQUE\dashboard\public\data'
os.makedirs(output_dir, exist_ok=True)

sheet_map = {}
for sn in wb.sheetnames:
    if 'Production' in sn: sheet_map[sn] = 'production'
    elif 'Comparatif' in sn: sheet_map[sn] = 'comparatif'
    elif 'Climat' in sn: sheet_map[sn] = 'climat'
    elif 'Commerce' in sn: sheet_map[sn] = 'commerce'
    elif 'curit' in sn: sheet_map[sn] = 'securite'
    elif 'Foncier' in sn: sheet_map[sn] = 'foncier'
    elif 'Finance' in sn: sheet_map[sn] = 'finance'
    elif 'Technologie' in sn: sheet_map[sn] = 'technologie'
    elif 'Sociaux' in sn: sheet_map[sn] = 'social'
    elif 'Durabilit' in sn: sheet_map[sn] = 'durabilite'
    elif 'Indices' in sn: sheet_map[sn] = 'indices'
    elif 'Statistiques' in sn: sheet_map[sn] = 'statistiques'
    elif 'lation' in sn: sheet_map[sn] = 'correlations'
    elif 'donn' in sn: sheet_map[sn] = 'metadonnees'

for sn, fname in sheet_map.items():
    ws = wb[sn]
    headers = [str(c.value) if c.value else f'col_{i}' for i, c in enumerate(ws[2])]
    data = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        rd = {}
        for i, val in enumerate(row):
            if val is not None and i < len(headers):
                if isinstance(val, float):
                    rd[headers[i]] = round(val, 4)
                else:
                    rd[headers[i]] = val if isinstance(val, int) else str(val)
        if rd:
            data.append(rd)
    with open(os.path.join(output_dir, f'{fname}.json'), 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)
    print(f'{fname}.json: {len(data)} records')

wb.close()
print('DONE')
